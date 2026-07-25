"""Shared 'CEO-ready' Excel formatting for MIS Studio exports.

One helper the Visual Builder, the lending templates, and the
Auto-Reporter all use, so every exported workbook looks the same:
bold title, styled header row, sensible column widths, thousands
separators, frozen header.
"""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="2E7BE6")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_TOTAL_FONT = Font(bold=True)


def write_formatted_sheet(
    writer: pd.ExcelWriter,
    frame: pd.DataFrame,
    sheet_name: str,
    *,
    title: str | None = None,
) -> None:
    """Write ``frame`` to ``sheet_name`` with the house formatting."""
    start_row = 2 if title else 0
    frame.to_excel(writer, sheet_name=sheet_name[:31], index=False, startrow=start_row)
    sheet = writer.sheets[sheet_name[:31]]

    if title:
        cell = sheet.cell(row=1, column=1, value=title)
        cell.font = _TITLE_FONT

    header_row = start_row + 1
    for col_index in range(1, len(frame.columns) + 1):
        cell = sheet.cell(row=header_row, column=col_index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    for col_index, column in enumerate(frame.columns, start=1):
        letter = get_column_letter(col_index)
        sample = frame[column].astype(str).head(50).map(len)
        width = max(len(str(column)), int(sample.max()) if len(sample) else 8) + 3
        sheet.column_dimensions[letter].width = min(max(width, 10), 42)
        if pd.api.types.is_numeric_dtype(frame[column]):
            for row_index in range(header_row + 1, header_row + 1 + len(frame)):
                sheet.cell(row=row_index, column=col_index).number_format = "#,##0.00"

    # Bold any TOTAL row (by convention the first cell says "TOTAL").
    for row_index in range(header_row + 1, header_row + 1 + len(frame)):
        if str(sheet.cell(row=row_index, column=1).value).strip().upper() == "TOTAL":
            for col_index in range(1, len(frame.columns) + 1):
                sheet.cell(row=row_index, column=col_index).font = _TOTAL_FONT

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
