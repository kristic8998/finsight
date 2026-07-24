"""MIS Generator: one click → CEO-ready Excel and HTML packs.

Daily / weekly / monthly variants share one builder; period affects the
KPI window and trend length. The Excel pack is a styled multi-sheet
workbook with an embedded trend chart; the HTML pack is a single
self-contained file that opens anywhere (and prints to PDF from any
browser — deliberate: no heavyweight PDF dependency).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from ..core.paths import reports_dir  # noqa: E402
from ..modules.executive import ExecutiveBrief, ExecutiveService  # noqa: E402

logger = logging.getLogger(__name__)

_PERIOD_DAYS = {"daily": 14, "weekly": 60, "monthly": 180}

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")


@dataclass
class MisOutput:
    excel_path: Path
    html_path: Path
    chart_path: Path
    brief: ExecutiveBrief


class MisGenerator:
    """Builds the report pack from an ExecutiveService."""

    def __init__(self, executive: ExecutiveService) -> None:
        self._executive = executive

    def generate(
        self,
        period: str = "daily",
        out_dir: Path | str | None = None,
        as_of: date | None = None,
    ) -> MisOutput:
        if period not in _PERIOD_DAYS:
            raise ValueError(f"period must be one of {sorted(_PERIOD_DAYS)}")
        today = as_of or date.today()
        target_dir = Path(out_dir) if out_dir is not None else reports_dir()
        target_dir.mkdir(parents=True, exist_ok=True)
        stem = f"{period}_mis_{today.isoformat()}"

        brief = self._executive.brief(today)
        trend = self._executive._data.daily_collections(  # noqa: SLF001 - same package
            days=_PERIOD_DAYS[period], as_of=today
        )

        chart_path = self._trend_chart(trend, target_dir / f"{stem}_trend.png", period)
        excel_path = self._excel_pack(
            brief, trend, chart_path, target_dir / f"{stem}.xlsx", period, today
        )
        html_path = self._html_pack(brief, target_dir / f"{stem}.html", period, today)
        logger.info("%s MIS generated: %s", period, excel_path.name)
        return MisOutput(
            excel_path=excel_path, html_path=html_path, chart_path=chart_path, brief=brief
        )

    # ---- pieces ---------------------------------------------------------------
    @staticmethod
    def _trend_chart(trend: pd.DataFrame, path: Path, period: str) -> Path:
        fig, ax = plt.subplots(figsize=(8, 3), dpi=120)
        ax.plot(pd.to_datetime(trend["date"]), trend["collected"], color="#2E7BE6", lw=1.6)
        ax.fill_between(
            pd.to_datetime(trend["date"]), trend["collected"], color="#2E7BE6", alpha=0.15
        )
        ax.set_title(f"Collections trend ({period} MIS window)")
        ax.set_ylabel("Collected")
        ax.grid(alpha=0.25)
        fig.autofmt_xdate()
        fig.tight_layout()
        fig.savefig(path)
        plt.close(fig)
        return path

    def _excel_pack(
        self,
        brief: ExecutiveBrief,
        trend: pd.DataFrame,
        chart: Path,
        path: Path,
        period: str,
        today: date,
    ) -> Path:
        kpis = brief.kpis
        kpi_frame = pd.DataFrame(
            {
                "KPI": [
                    "Business Health Score",
                    "Portfolio Outstanding",
                    "Active Loans",
                    "Disbursed (MTD)",
                    "Due (MTD)",
                    "Collected (MTD)",
                    "Collection Efficiency %",
                    "Target Achievement %",
                    "Overdue Amount",
                    "PAR %",
                    "NPA Loans",
                    "NPA Amount",
                    "Interest Collected (MTD)",
                    "Growth MoM %",
                ],
                "Value": [
                    f"{brief.health.score:.0f} ({brief.health.grade})",
                    kpis.portfolio_outstanding,
                    kpis.active_loans,
                    kpis.disbursed_mtd,
                    kpis.due_mtd,
                    kpis.collected_mtd,
                    kpis.efficiency_mtd,
                    kpis.target_achievement,
                    kpis.overdue_amount,
                    kpis.par_pct,
                    kpis.npa_loans,
                    kpis.npa_amount,
                    kpis.interest_collected_mtd,
                    kpis.growth_mom_pct,
                ],
            }
        )
        insights_frame = pd.DataFrame(
            [
                {
                    "severity": i.severity,
                    "finding": i.title,
                    "detail": i.detail,
                    "recommendation": i.recommendation,
                }
                for i in brief.insights
            ]
        )

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            kpi_frame.to_excel(writer, sheet_name="Summary", index=False, startrow=3)
            brief.branches.to_excel(writer, sheet_name="Branches", index=False)
            brief.dpd.to_excel(writer, sheet_name="DPD Buckets", index=False)
            brief.product_mix.to_excel(writer, sheet_name="Product Mix", index=False)
            insights_frame.to_excel(writer, sheet_name="Insights", index=False)
            trend.to_excel(writer, sheet_name="Trend Data", index=False)

            summary = writer.sheets["Summary"]
            summary["A1"] = f"{period.title()} MIS — {today.isoformat()}"
            summary["A1"].font = _TITLE_FONT
            summary["A2"] = brief.summary_text
            summary["A2"].alignment = Alignment(wrap_text=True)
            summary.merge_cells("A2:F2")
            summary.row_dimensions[2].height = 42

            for sheet_name in writer.sheets:
                sheet = writer.sheets[sheet_name]
                header_row = 4 if sheet_name == "Summary" else 1
                for cell in sheet[header_row]:
                    if cell.value is not None:
                        cell.fill = _HEADER_FILL
                        cell.font = _HEADER_FONT
                sheet.freeze_panes = f"A{header_row + 1}"
                for idx, column_cells in enumerate(sheet.columns, start=1):
                    width = max(
                        (len(str(c.value)) for c in column_cells if c.value is not None),
                        default=8,
                    )
                    sheet.column_dimensions[get_column_letter(idx)].width = min(width + 2, 60)

            from openpyxl.drawing.image import Image as XlImage

            img = XlImage(str(chart))
            writer.sheets["Summary"].add_image(img, "H4")
        return path

    @staticmethod
    def _severity_color(severity: str) -> str:
        return {"alert": "#C62828", "watch": "#EF6C00", "good": "#2E7D32"}[severity]

    def _html_pack(self, brief: ExecutiveBrief, path: Path, period: str, today: date) -> Path:
        kpis = brief.kpis
        cards = [
            ("Health", f"{brief.health.score:.0f} ({brief.health.grade})"),
            ("Portfolio", f"₹{kpis.portfolio_outstanding:,.0f}"),
            ("Collected MTD", f"₹{kpis.collected_mtd:,.0f}"),
            ("Efficiency", f"{kpis.efficiency_mtd:.0f}%"),
            ("Overdue", f"₹{kpis.overdue_amount:,.0f}"),
            ("PAR", f"{kpis.par_pct:.1f}%"),
            ("Active Loans", f"{kpis.active_loans:,}"),
            ("Growth MoM", f"{kpis.growth_mom_pct:+.1f}%"),
        ]
        card_html = "".join(
            f'<div class="card"><div class="k">{label}</div><div class="v">{value}</div></div>'
            for label, value in cards
        )
        insight_html = "".join(
            f'<div class="insight" style="border-left-color:{self._severity_color(i.severity)}">'
            f"<b>{i.title}</b><br><span>{i.detail}</span><br>"
            f"<i>→ {i.recommendation}</i></div>"
            for i in brief.insights
        )
        branch_table = brief.branches.to_html(index=False, border=0)
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>{period.title()} MIS {today.isoformat()}</title><style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;color:#1c2733;background:#f6f8fb}}
h1{{color:#1F4E79;margin-bottom:2px}} .sub{{color:#5a6b7b;margin-bottom:18px}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px}}
.card{{background:#fff;border-radius:10px;padding:12px 18px;
box-shadow:0 1px 4px #0002;min-width:130px}}
.card .k{{font-size:12px;color:#5a6b7b}} .card .v{{font-size:20px;font-weight:600}}
.insight{{background:#fff;border-left:5px solid;border-radius:8px;padding:10px 14px;margin:8px 0;
box-shadow:0 1px 3px #0001}} .insight span{{color:#44535f}} .insight i{{color:#1F4E79}}
table{{border-collapse:collapse;background:#fff;box-shadow:0 1px 4px #0002;border-radius:8px}}
th{{background:#1F4E79;color:#fff;padding:7px 12px;text-align:left;font-size:13px}}
td{{padding:6px 12px;border-bottom:1px solid #e8edf3;font-size:13px}}
@media print{{body{{background:#fff}}}}</style></head><body>
<h1>{period.title()} MIS — {today.isoformat()}</h1>
<div class="sub">{brief.summary_text}</div>
<div class="cards">{card_html}</div>
<h2>Executive insights</h2>{insight_html}
<h2>Branch ranking</h2>{branch_table}
<p class="sub">Generated by FinSight. Print → Save as PDF for a board copy.</p>
</body></html>"""
        path.write_text(html, encoding="utf-8")
        return path
