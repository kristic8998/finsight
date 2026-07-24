"""MIS report catalog + custom template engine.

The named business reports a lending analyst produces every day —
Collections, Disbursement, Recovery/Overdue, Branch, Employee, Product —
each built from the same LendingDataService read-model the dashboard
uses (so a branch number in the MIS always equals the dashboard).
Custom templates let the analyst compose their own multi-sheet pack
from named sections and reuse it forever; templates persist in the
app database.

Outputs: styled multi-sheet .xlsx (chunk-written via openpyxl) or .csv
(one file per section).
"""

from __future__ import annotations

import json
import logging
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.appdb import AppDB
from ..core.paths import reports_dir
from ..data.queries import LendingDataService

logger = logging.getLogger(__name__)

_TEMPLATES_KEY = "mis_templates"
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13, color="1F4E79")


class CatalogError(Exception):
    """Raised for unknown reports/sections or unusable template input."""


@dataclass
class CatalogOutput:
    """Where a generated report landed."""

    report: str
    paths: list[Path]
    sheets: dict[str, int]  # sheet/section name -> rows

    @property
    def summary(self) -> str:
        parts = ", ".join(f"{name} ({rows})" for name, rows in self.sheets.items())
        return f"{self.report}: {parts} → {self.paths[0].name}"


class MisCatalog:
    """Named business reports + user-defined templates."""

    def __init__(self, data: LendingDataService, appdb: AppDB, company_name: str = "") -> None:
        self._data = data
        self._appdb = appdb
        self._company = company_name.strip()

    # ------------------------------------------------------------------ sections
    # Every section is a pure builder returning a DataFrame. Reports and
    # custom templates are just named selections of sections.
    def _sec_collections_by_day(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.daily_collections(days=days, as_of=as_of)

    def _paid_window(self, days: int, as_of: date) -> pd.DataFrame:
        pays = self._data.payments_status(as_of)
        paid = pays[pays["paid_date"].notna()].copy()
        paid["paid_on"] = pd.to_datetime(paid["paid_date"]).dt.date
        start = as_of - timedelta(days=days)
        return paid[(paid["paid_on"] >= start) & (paid["paid_on"] <= as_of)]

    def _sec_collections_by_branch(self, days: int, as_of: date) -> pd.DataFrame:
        window = self._paid_window(days, as_of)
        out = (
            window.groupby("branch")["amount_paid"]
            .agg(collected="sum", receipts="count")
            .reset_index()
            .sort_values("collected", ascending=False)
        )
        return out.round(2)

    def _sec_collections_by_mode(self, days: int, as_of: date) -> pd.DataFrame:
        window = self._paid_window(days, as_of)
        out = (
            window.groupby("mode")["amount_paid"]
            .agg(collected="sum", receipts="count")
            .reset_index()
            .sort_values("collected", ascending=False)
        )
        return out.round(2)

    def _disbursed_window(self, days: int, as_of: date) -> pd.DataFrame:
        loans = self._data.loans()
        loans["disbursed"] = pd.to_datetime(loans["disbursed_on"]).dt.date
        start = as_of - timedelta(days=days)
        return loans[(loans["disbursed"] >= start) & (loans["disbursed"] <= as_of)]

    def _sec_disbursement_by_branch(self, days: int, as_of: date) -> pd.DataFrame:
        window = self._disbursed_window(days, as_of)
        out = (
            window.groupby("branch")["principal"]
            .agg(disbursed="sum", loans="count")
            .reset_index()
            .sort_values("disbursed", ascending=False)
        )
        out["avg_ticket"] = (out["disbursed"] / out["loans"].replace(0, pd.NA)).astype(float)
        return out.round(2)

    def _sec_disbursement_by_product(self, days: int, as_of: date) -> pd.DataFrame:
        window = self._disbursed_window(days, as_of)
        out = (
            window.groupby("product")["principal"]
            .agg(disbursed="sum", loans="count")
            .reset_index()
            .sort_values("disbursed", ascending=False)
        )
        return out.round(2)

    def _sec_dpd_buckets(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.dpd_buckets(as_of)

    def _sec_top_overdue(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.overdue_loans(limit=200)

    def _sec_branch_ranking(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.branch_summary(as_of)

    def _sec_region_rollup(self, days: int, as_of: date) -> pd.DataFrame:
        loans = self._data.loans()
        active = loans[loans["status"] != "closed"]
        out = (
            active.groupby("region")["principal"]
            .agg(principal="sum", loans="count")
            .reset_index()
            .sort_values("principal", ascending=False)
        )
        return out.round(2)

    def _sec_officer_productivity(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.officer_productivity(as_of)

    def _sec_product_mix(self, days: int, as_of: date) -> pd.DataFrame:
        return self._data.product_mix()

    def _sec_product_by_branch(self, days: int, as_of: date) -> pd.DataFrame:
        loans = self._data.loans()
        active = loans[loans["status"] != "closed"]
        pivot = pd.pivot_table(
            active,
            values="principal",
            index="branch",
            columns="product",
            aggfunc="sum",
            fill_value=0.0,
        ).reset_index()
        return pivot.round(0)

    def _sec_kpi_summary(self, days: int, as_of: date) -> pd.DataFrame:
        kpis = self._data.kpis(as_of)
        return pd.DataFrame(
            {
                "kpi": [
                    "portfolio_outstanding",
                    "active_loans",
                    "npa_loans",
                    "npa_amount",
                    "disbursed_mtd",
                    "due_mtd",
                    "collected_mtd",
                    "efficiency_mtd_pct",
                    "target_achievement_pct",
                    "overdue_amount",
                    "par_pct",
                    "growth_mom_pct",
                ],
                "value": [
                    kpis.portfolio_outstanding,
                    kpis.active_loans,
                    kpis.npa_loans,
                    kpis.npa_amount,
                    kpis.disbursed_mtd,
                    kpis.due_mtd,
                    kpis.collected_mtd,
                    kpis.efficiency_mtd,
                    kpis.target_achievement,
                    kpis.overdue_amount,
                    kpis.par_pct,
                    kpis.growth_mom_pct,
                ],
            }
        )

    # ------------------------------------------------------------------ registry
    @property
    def sections(self) -> dict[str, tuple[str, Callable[[int, date], pd.DataFrame]]]:
        """section id -> (sheet title, builder)."""
        return {
            "kpi_summary": ("KPI Summary", self._sec_kpi_summary),
            "collections_by_day": ("Collections by Day", self._sec_collections_by_day),
            "collections_by_branch": ("Collections by Branch", self._sec_collections_by_branch),
            "collections_by_mode": ("Collections by Mode", self._sec_collections_by_mode),
            "disbursement_by_branch": ("Disbursement by Branch", self._sec_disbursement_by_branch),
            "disbursement_by_product": (
                "Disbursement by Product",
                self._sec_disbursement_by_product,
            ),
            "dpd_buckets": ("DPD Buckets", self._sec_dpd_buckets),
            "top_overdue": ("Top Overdue Loans", self._sec_top_overdue),
            "branch_ranking": ("Branch Ranking", self._sec_branch_ranking),
            "region_rollup": ("Region Rollup", self._sec_region_rollup),
            "officer_productivity": ("Officer Productivity", self._sec_officer_productivity),
            "product_mix": ("Product Mix", self._sec_product_mix),
            "product_by_branch": ("Product x Branch", self._sec_product_by_branch),
        }

    REPORTS: dict[str, tuple[str, tuple[str, ...]]] = {
        "collections": (
            "Collections MIS",
            ("kpi_summary", "collections_by_day", "collections_by_branch", "collections_by_mode"),
        ),
        "disbursement": (
            "Disbursement MIS",
            ("kpi_summary", "disbursement_by_branch", "disbursement_by_product"),
        ),
        "recovery": (
            "Recovery & Overdue MIS",
            ("kpi_summary", "dpd_buckets", "branch_ranking", "top_overdue"),
        ),
        "branch": ("Branch MIS", ("branch_ranking", "region_rollup", "product_by_branch")),
        "employee": ("Employee MIS", ("officer_productivity",)),
        "product": ("Product MIS", ("product_mix", "product_by_branch", "disbursement_by_product")),
    }

    def report_names(self) -> dict[str, str]:
        return {rid: title for rid, (title, _s) in self.REPORTS.items()}

    # ------------------------------------------------------------------ build
    def build(
        self,
        report_id: str,
        days: int = 30,
        as_of: date | None = None,
        fmt: str = "xlsx",
        out_dir: Path | str | None = None,
    ) -> CatalogOutput:
        """Generate a catalog report as xlsx (multi-sheet) or csv (per section)."""
        if report_id not in self.REPORTS:
            raise CatalogError(
                f"unknown report '{report_id}' " f"(have: {', '.join(self.REPORTS)})"
            )
        title, section_ids = self.REPORTS[report_id]
        return self._build_sections(title, section_ids, days, as_of, fmt, out_dir)

    def _build_sections(
        self,
        title: str,
        section_ids: tuple[str, ...] | list[str],
        days: int,
        as_of: date | None,
        fmt: str,
        out_dir: Path | str | None,
    ) -> CatalogOutput:
        if fmt not in ("xlsx", "csv"):
            raise CatalogError("format must be 'xlsx' or 'csv'")
        today = as_of or date.today()
        target = Path(out_dir) if out_dir is not None else reports_dir()
        target.mkdir(parents=True, exist_ok=True)

        frames: dict[str, pd.DataFrame] = {}
        registry = self.sections
        for section_id in section_ids:
            if section_id not in registry:
                raise CatalogError(f"unknown section '{section_id}'")
            sheet_title, builder = registry[section_id]
            frames[sheet_title] = builder(days, today)

        stem = f"{title.lower().replace(' ', '_').replace('&', 'and')}_{today.isoformat()}"
        paths: list[Path] = []
        if fmt == "xlsx":
            path = target / f"{stem}.xlsx"
            self._write_workbook(frames, path, title, today, days)
            paths.append(path)
        else:
            for sheet_title, frame in frames.items():
                safe = sheet_title.lower().replace(" ", "_").replace("x", "x")
                path = target / f"{stem}__{safe}.csv"
                frame.to_csv(path, index=False)
                paths.append(path)

        output = CatalogOutput(
            report=title,
            paths=paths,
            sheets={name: int(len(f)) for name, f in frames.items()},
        )
        logger.info("catalog report generated: %s", output.summary)
        return output

    def _write_workbook(
        self,
        frames: dict[str, pd.DataFrame],
        path: Path,
        title: str,
        today: date,
        days: int,
    ) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_title, frame in frames.items():
                sheet_name = sheet_title[:31]
                frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                sheet = writer.sheets[sheet_name]
                header = f"{title} — {sheet_title} — {today.isoformat()} (window {days}d)"
                if self._company:
                    header = f"{self._company} · {header}"
                sheet["A1"] = header
                sheet["A1"].font = _TITLE_FONT
                for cell in sheet[3]:
                    if cell.value is not None:
                        cell.fill = _HEADER_FILL
                        cell.font = _HEADER_FONT
                sheet.freeze_panes = "A4"
                for idx, column_cells in enumerate(sheet.columns, start=1):
                    width = max(
                        (len(str(c.value)) for c in column_cells if c.value is not None),
                        default=8,
                    )
                    sheet.column_dimensions[get_column_letter(idx)].width = min(width + 2, 42)

    # ------------------------------------------------------------------ templates
    def templates(self) -> dict[str, list[str]]:
        """name -> ordered section ids."""
        raw = self._appdb.get_value(_TEMPLATES_KEY, "{}")
        data = json.loads(raw) if isinstance(raw, str) else raw
        return {str(k): list(v) for k, v in dict(data).items()}

    def save_template(self, name: str, section_ids: list[str]) -> None:
        name = name.strip()
        if not name:
            raise CatalogError("template needs a name")
        unknown = [s for s in section_ids if s not in self.sections]
        if unknown:
            raise CatalogError(f"unknown section(s): {', '.join(unknown)}")
        if not section_ids:
            raise CatalogError("pick at least one section")
        current = self.templates()
        current[name] = list(section_ids)
        self._appdb.set_value(_TEMPLATES_KEY, current)

    def delete_template(self, name: str) -> None:
        current = self.templates()
        current.pop(name, None)
        self._appdb.set_value(_TEMPLATES_KEY, current)

    def build_from_template(
        self,
        name: str,
        days: int = 30,
        as_of: date | None = None,
        fmt: str = "xlsx",
        out_dir: Path | str | None = None,
    ) -> CatalogOutput:
        templates = self.templates()
        if name not in templates:
            raise CatalogError(f"unknown template '{name}'")
        return self._build_sections(name, templates[name], days, as_of, fmt, out_dir)


def recent_reports(limit: int = 12) -> list[Path]:
    """Most recently generated report files, newest first."""
    files = [
        p
        for p in reports_dir().iterdir()
        if p.is_file() and p.suffix.lower() in (".xlsx", ".html", ".csv")
    ]
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:limit]
